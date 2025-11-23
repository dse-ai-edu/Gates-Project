import random
import json
import io
from flask import request, jsonify
from openpyxl import load_workbook
from app import app


##IMPLEMENT session save for use of file data later
@app.route('/upload', methods=['POST'])
def upload_file():
    # Ensure a file was sent
    if 'file' not in request.files:
        return jsonify({"error": "No file path"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    valid_types = {
        'text/plain': process_text,
        'application/json': process_json,
        'text/csv': process_csv,
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': process_xlsx
    }

    if file.content_type not in valid_types:
        return jsonify({"error": "Invalid file type"}), 400

    content = file.read()
    random_flag = request.form.get('random') == 'true'
    processor = valid_types[file.content_type]

    result, error = processor(content, random_flag)

    if error:
        return jsonify({"error": error}), 400

    return jsonify({"content": result})

def process_text(content, random_flag):
    try:
        content = content.decode('utf-8')
        if random_flag:
            lines = content.splitlines()
            return (random.choice(lines) if lines else ""), None
        return content, None
    except Exception as e:
        return None, f"Error processing text file: {str(e)}"

def process_csv(content, random_flag):
    try:
        content = content.decode('utf-8')
        rows = [row.split(',') for row in content.strip().splitlines()]
        if random_flag:
            if rows:
                return "\t".join(random.choice(rows)), None
            return "", None
        formatted_rows = ["\t".join(row) for row in rows]
        return "\n".join(formatted_rows), None
    except Exception as e:
        return None, f"Error processing CSV file: {str(e)}"

def process_json(content, random_flag):
    try:
        content = content.decode('utf-8')
        data = json.loads(content)
    except Exception as e:
        return None, f"Invalid JSON file: {str(e)}"

    if random_flag:
        if isinstance(data, list):
            return json.dumps(random.choice(data)), None
        elif isinstance(data, dict):
            keys = list(data.keys())
            if keys:
                random_key = random.choice(keys)
                return f"{random_key}: {data[random_key]}", None
            return "", None
        return str(data), None
    return json.dumps(data, indent=2), None

def process_xlsx(content, random_flag):
    try:
        file_stream = io.BytesIO(content)
        workbook = load_workbook(file_stream, data_only=True)
        sheet = workbook.active
        rows = ["\t".join(str(cell) if cell is not None else "" for cell in row) for row in sheet.iter_rows(values_only=True)]
        if random_flag:
            return (random.choice(rows), None) if rows else ("", None)
        return "\n".join(rows), None
    except Exception as e:
        return None, f"Error processing XLSX file: {str(e)}"